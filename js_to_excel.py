# !/usr/bin/env python
# encoding=utf-8
# author: zhanzq
# email : zhanzhiqiang09@126.com 
# date  : 2022/3/11
#

import re
import os
import time
from functools import wraps

import traceback
import openpyxl
from openpyxl import Workbook


def time_cost(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        start_time = time.time()
        ret_res = f(*args, **kwargs)
        end_time = time.time()
        param_lst = []
        for arg in args:
            param_lst.append(_convert_to_str(arg))

        for key, val in kwargs.items():
            param_lst.append("{}={}".format(key, _convert_to_str(val)))
        params = ", ".join(param_lst)
        print("{}({}) cost time: {:.3f} seconds".format(f.__name__, params, end_time - start_time))
        return ret_res

    return decorated


def save_jsons_into_sheet(wb, json_lst, col_name_lst=None, sheet_name="Title", overwrite=False):
    """
    save json list into sheet of xlsx file
    :param wb: workbook, fp of xlsx file
    :param json_lst: the data to store
    :param col_name_lst: the column names of the sheet data, ordered according to col_name_lst
    :param sheet_name: the sheet name to store given data
    :param overwrite: overwrite the old data or not, default "False"
    :return:
    """
    if col_name_lst is None:
        col_name_lst = list(json_lst[0].keys())

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if overwrite:
            wb.remove(sheet)
            sheet = wb.create_sheet(title=sheet_name)
    else:
        sheet = wb.create_sheet(title=sheet_name)

    # write first line with col_names
    for j, col_name in enumerate(col_name_lst):
        sheet.cell(row=1, column=j + 1).value = col_name

    # write json obj
    col_sz = len(col_name_lst)
    for i, obj in enumerate(json_lst):
        for j in range(col_sz):
            val = obj.get(col_name_lst[j])
            if val is None:
                val = ""
            sheet.cell(row=i + 2, column=j + 1).value = str(val)


@time_cost
def save_jsons_into_xlsx(json_lst, xlsx_path, col_name_lst=None, sheet_name: str = "Title", overwrite=True):
    """
    write json list into xlsx file, treate keys as the column names
    :param json_lst: data to store with json format
    :param xlsx_path: data path
    :param col_name_lst: the column names of the sheet data, ordered according to col_name_lst
    :param sheet_name: the sheet name to store given data
    :param overwrite: overwrite the old data or not, default "True"
    :return:
    """
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.get_sheet_by_name("Sheet"))

    save_jsons_into_sheet(wb, json_lst=json_lst, col_name_lst=col_name_lst, sheet_name=sheet_name, overwrite=overwrite)

    wb.save(xlsx_path)


@time_cost
def load_jsons_from_xlsx(xlsx_path, sheet_name=None):
    """
    read data from xlsx file, and format it in dict, keys are all the sheet names or the give sheet_name
    :param xlsx_path: data file path
    :param sheet_name: the sheet name from which to read data
    :return:
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        json_dct = {}
        sheet_names = wb.sheetnames
        if sheet_name and sheet_name in sheet_names:
            sheet_names = [sheet_name]
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            json_dct[sheet_name] = load_jsons_from_sheet_v2(sheet)

        return json_dct
    except FileNotFoundError as e:
        print(e)
        return None


def eval_cell(val):
    try:
        val = eval(val)
    except NameError:
        pass
    finally:
        return val


def load_jsons_from_sheet(sheet):
    """
    read data from sheet of xlsx file, and format it in json list
    :param sheet: the sheet name from which to read data
    :return:
    """
    rows = sheet.rows
    keys = []
    try:
        keys = [it.value for it in rows.__next__() if it.value]
    except StopIteration:
        pass

    json_lst = []
    for row in rows:
        values = [eval_cell(it.value) for it in row]
        obj = {key: val for key, val in zip(keys, values)}
        json_lst.append(obj)

    return json_lst


def convert_pos_26bit(s):
    """
    convert A2 to (2, 1)
    convert AA2 to (2, 27)
    :param s:
    :return:
    """
    first = 0
    second = 0
    for ch in s:
        if 'A' <= ch <= 'Z':
            second *= 26
            second += ord(ch) - ord('A') + 1
        else:
            first *= 10
            first += ord(ch) - ord('0')

    return first, second


def convert_str_to_cell_range(s):
    """
    convert "A2:A8" to [[(2, 1)], [(3, 1)], [(4, 1)], [(5, 1)], [(6, 1)], [(7, 1)], [(8, 1)]]
    convert "A2:c8" to [[(2, 1), (2, 2), (2, 3)],
                       [(3, 1), (3, 2), (3, 3)],
                       [(4, 1), (4, 2), (4, 3)],
                       [(5, 1), (5, 2), (5, 3)],
                       [(6, 1), (6, 2), (6, 3)],
                       [(7, 1), (7, 2), (7, 3)],
                       [(8, 1), (8, 2), (8, 3)]]
    :param s:
    :return:
    """
    items = s.split(":")
    first1, second1 = convert_pos_26bit(items[0])
    first2, second2 = convert_pos_26bit(items[1])

    cell_range = []
    for i in range(first1, first2 + 1):
        row = []
        for j in range(second1, second2 + 1):
            row.append((i, j))
        cell_range.append(row)

    return cell_range


def get_merged_cells_from_sheet(sheet):
    mp = {}
    if openpyxl.__version__ >= "3.0.1":
        range_lst = sheet.merged_cells.ranges
        for _range in range_lst:
            pos_lst = list(_range.rows)
            i, j = pos_lst[0][0]
            val = sheet[i][j - 1].value  # xlxs row: starts from 0, column: starts from 1
            mp[val] = []
            for lst in pos_lst:
                mp[val].extend(lst)

        return mp
    else:
        range_lst = sheet.merged_cell_ranges
        for range_s in range_lst:
            pos_lst = convert_str_to_cell_range(range_s)
            i, j = pos_lst[0][0]
            val = sheet[i][j - 1].value  # xlxs row: starts from 0, column: starts from 1
            mp[val] = []
            for lst in pos_lst:
                mp[val].extend(lst)

        return mp


def load_jsons_from_sheet_v2(sheet):
    rows = sheet.rows
    col_names = [it.value for it in rows.__next__()]
    lst = []
    for row in rows:
        values = [eval_cell(it.value) for it in row]
        lst.append(values)

    merged_cells = get_merged_cells_from_sheet(sheet)
    for val, pos_lst in merged_cells.items():
        for i, j in pos_lst:
            lst[i - 2][j - 1] = val

    json_lst = []
    for item in lst:
        json_lst.append(dict(zip(col_names, item)))

    return json_lst


def _convert_to_str(arg):
    if type(arg) is str:
        return arg
    elif type(arg) is int or type(arg) is float or type(arg) is bool:
        return str(arg)
    elif callable(arg):
        return arg.__name__
    else:
        return type(arg).__name__


# 获取antlr中所有可能的预定义方法
def get_candidate_functions(code):
    if not code:
        return set()
    function_ptn = "(\w+)\("
    return set(re.findall(function_ptn, code))


def get_all_pre_defined_functions(tpl_xlsx_path):
    json_dct = load_jsons_from_xlsx(tpl_xlsx_path)
    st = set()
    for car_type, json_lst in json_dct.items():
        for item in json_lst:
            tmp = get_candidate_functions(item["antlr"])
            st = st.union(tmp)

    for key in st:
        print("\"function {}() ".format(key) + "{ }\",")


def get_all_pre_defined_functions(tpl_xlsx_path):
    json_dct = load_jsons_from_xlsx(tpl_xlsx_path)
    st = set()
    for car_type, json_lst in json_dct.items():
        for item in json_lst:
            tmp = get_candidate_functions(item["antlr"])
            st = st.union(tmp)

    for key in st:
        print("\"function {}() ".format(key) + "{ }\",")


# all methods in js function <---> json object
def get_pre_defined_function():
    function_lst = [
        "function data(key) { }",
        "function param(name) { }",
        "function delay(time) { }",
        "function cmd(command) { }",
        "function tts(text) { }",
        "function eqMethod(left, right) { }",
        "function neqMethod(left, right) { }",
        "function ltMethod(left, right) { }",
        "function geMethod(left, right) { }",
        "function gtMethod(left, right) { }",
        "function and(left, right) { }",
        "function or(left, right) { }",
        "function int(num) { }",
        "function isSubscribe(key) { }",
        "function assert(state) { }",
    ]

    return function_lst


def save_js_functions(js_functions, output_path):
    out_lines = [it + "\n" for it in js_functions]
    with open(output_path, "a") as writer:
        writer.writelines(out_lines)


def convert_var_to_js_grammar(var_statement):
    """
    convert var statement from antlr to js grammar
    :param var_statement: var statement in antlr grammar,
    when parsing, it must fullmatch "var(FIRST,SECOND);",
    otherwise, return the original statement
    :return: str, in js grammar, like "var FIRST = SECOND;"
    """
    statement = var_statement
    lspace_num = len(statement) - len(statement.lstrip())
    var_ptn = "var *\( *(\w+), *(.*)\) *;"
    if re.fullmatch(var_ptn, statement.strip()):
        first, second = re.findall(var_ptn, statement)[0]
        return "{}var {} = {};".format(" " * lspace_num, first, second)
    else:
        return var_statement


def revert_var_from_js_grammar(var_statement):
    """
    revert var statement from js to antlr grammar
    :param var_statement: var statement in js grammar,
    when parsing, it must fullmatch "var FIRST = SECOND;",
    otherwise, return the original statement
    :return: str, in antlr grammar, like "var(FIRST,SECOND);"
    """
    statement = var_statement
    lspace_num = len(statement) - len(statement.lstrip())
    var_ptn = "var (.*) = (.*);"
    if re.fullmatch(var_ptn, statement.strip()):
        items = re.findall(var_ptn, statement.strip())[0]
        first, second = items

        return " " * lspace_num + "var({},{});".format(first, second)
    else:
        return var_statement


def convert_elif_to_js_grammar(line):
    return line.replace("elif", "else if")


def revert_elif_from_js_grammar(line):
    return line.replace("else if", "elif")


def convert_antlr_to_js(antlr, leading_space=" " * 4):
    if not antlr:
        return ""

    antlr_codes = antlr.split("\n")
    js_codes = []
    if len(antlr_codes) == 1:
        js_codes.append(antlr_codes[0])
    else:
        for line in antlr_codes:
            if line.lstrip().startswith("var("):
                js_codes.append(convert_var_to_js_grammar(line))
            elif "elif" in line:
                js_codes.append(convert_elif_to_js_grammar(line))
            else:
                js_codes.append(line)

    js_codes = "\n".join([leading_space + line for line in js_codes])

    return js_codes


def revert_antlr_from_js(js_codes, leading_spaces=" " * 4):
    if not js_codes:
        return ""

    antlr_codes = []
    js_codes = js_codes.split("\n")
    if len(js_codes) == 1:
        return js_codes[0]
    else:
        for line in js_codes:
            if line.lstrip().startswith("var "):
                antlr_codes.append(revert_var_from_js_grammar(line))
            elif "else if" in line:
                antlr_codes.append(revert_elif_from_js_grammar(line))
            else:
                antlr_codes.append(line)

        return "\n".join(antlr_codes)


def convert_tpl_item_to_js_function(tpl_item):
    """
    convert tpl_item to js function
    :param tpl_item: input tpl item, in format {"intent": INTENT, "param": PARAM, "command": CMD, "antlr": TPL}
    :return: str, represent the corresponding js codes
for examle:
input:
{'intent': 'charge_port_open',
 'param': None,
 'command': None,
 'antlr': "if(isSubscribe('charge.direct.port.on.status')){\n var(_status,data('charge.direct.port.on.status'));\n if(eqMethod(_status,1)){\n  cmd('charge.direct.port.on');\n  tts('充电口盖已打开');\n }\n elif(eqMethod(_status,2)){\n  tts('行车过程中,充电口盖暂不支持打开哦');\n }\n elif(eqMethod(_status,3)){\n  tts('当前充电口状态不支持打开或关闭哦');\n }\n}else{\n assert(isSubscribe('charge.direct.port.on'));\n if(data('charge.direct.port.on.support')){\n  cmd('charge.direct.port.on');\n  tts('充电口盖已打开');\n }else{\n  tts('行车过程中或充电枪插入或充电口盖打开时，充电口盖暂不支持打开哦');\n }\n}
}
output:
'function antlr(intent="charge_port_open", param="None", command="None") {\n    if(isSubscribe(\'charge.direct.port.on.status\')){\n     var _status = data(\'charge.direct.port.on.status\');\n     if(eqMethod(_status,1)){\n      cmd(\'charge.direct.port.on\');\n      tts(\'充电口盖已打开\');\n     }\n     else if(eqMethod(_status,2)){\n      tts(\'行车过程中,充电口盖暂不支持打开哦\');\n     }\n     else if(eqMethod(_status,3)){\n      tts(\'当前充电口状态不支持打开或关闭哦\');\n     }\n    }else{\n     assert(isSubscribe(\'charge.direct.port.on\'));\n     if(data(\'charge.direct.port.on.support\')){\n      cmd(\'charge.direct.port.on\');\n      tts(\'充电口盖已打开\');\n     }else{\n      tts(\'行车过程中或充电枪插入或充电口盖打开时，充电口盖暂不支持打开哦\');\n     }\n    }\n}'
    """
    try:
        intent = tpl_item["intent"]
        param = tpl_item["param"]
        command = tpl_item["command"]
        antlr = tpl_item["antlr"]
        js_head = 'function antlr(intent="{}", param="{}", command="{}")'.format(intent or "None", param or "None",
                                                                                 command or "None") + " {"
        js_codes = convert_antlr_to_js(antlr=antlr)
        js_tail = "}"
        js_func = "\n".join([js_head, js_codes, js_tail])

        return js_func
    except:
        print(tpl_item)
        traceback.print_exc()

        return ""


def revert_tpl_item_from_js_function(js_func):
    js_func = js_func.strip().split("\n")

    js_head, js_codes, js_tail = js_func[0], "\n".join(js_func[1:-1]), js_func[-1]
    params = re.findall('\"(.*?)\"', js_head)
    tpl_item = {}
    for key, val in zip(["intent", "param", "command"], params):
        tpl_item[key] = None if val == "None" else val

    antlr = revert_antlr_from_js(js_codes)
    tpl_item["antlr"] = antlr

    return tpl_item


@time_cost
def save_jsons_to_js_file(json_lst, js_path, overwrite=True):
    if overwrite and os.path.exists(js_path):
        os.remove(js_path)
    save_js_functions(get_pre_defined_function(), output_path=js_path)

    js_functions = []
    for tpl_item in json_lst:
        js_func = convert_tpl_item_to_js_function(tpl_item=tpl_item)
        js_functions.append(js_func)

    save_js_functions(js_functions, output_path=js_path)
    return


@time_cost
def load_jsons_from_js_file(js_path):
    json_lst = []
    with open(js_path, "r") as reader:
        lines = "".join(reader.readlines())
        js_functions = re.split(pattern="function antlr", string=lines)[1:]  # ignore predefined functions
        for js_func in js_functions:
            tpl_item = revert_tpl_item_from_js_function(js_func)
            json_lst.append(tpl_item)

    return json_lst


@time_cost
def convert_js_to_xlsx(js_path, xlsx_path, sheet_name: str=None):
    if not sheet_name:
        assert os.path.isdir(js_path), "when sheet_name is not specified, js_path should be dir"
        fnames = os.listdir(js_path)
        for fname in fnames:
            sheet_name = str(fname.split("/")[-1]).split(".")[0]
            js_path_i = os.path.join(js_path, fname)
            json_lst = load_jsons_from_js_file(js_path_i)
            save_jsons_into_xlsx(xlsx_path=xlsx_path, json_lst=json_lst, sheet_name=str(sheet_name))
    else:
        assert os.path.isfile(js_path), "when sheet_name is specified, js_path should be file"
        json_lst = load_jsons_from_js_file(js_path)
        save_jsons_into_xlsx(xlsx_path=xlsx_path, json_lst=json_lst, sheet_name=sheet_name)

    return

@time_cost
def revert_js_from_xlsx(js_path, xlsx_path, sheet_name=None):
    if not sheet_name:
        if os.path.exists(js_path):
            assert os.path.isdir(js_path), "when sheet_name is not specified, js_path should be dir"
        else:
            os.mkdir(js_path)
        json_dct = load_jsons_from_xlsx(xlsx_path=xlsx_path, sheet_name=sheet_name)
        for sheet_name, json_lst in json_dct.items():
            js_path_i = os.path.join(js_path, "{}.js".format(sheet_name))
            save_jsons_to_js_file(js_path=js_path_i, json_lst=json_lst)
    else:
        assert os.path.isfile(js_path), "when sheet_name is specified, js_path should be file"
        json_lst = load_jsons_from_xlsx(xlsx_path=xlsx_path, sheet_name=sheet_name)[sheet_name]
        save_jsons_to_js_file(js_path=js_path, json_lst=json_lst)

    return


def format_string(s, length=100):
    """
    format the print info of given string *s*
    :param s: which to format
    :param length: the output length, default 100
    :return: with format as follows:
    "*********************    hello, world!    **********************"
    """
    res = "   {}   ".format(s)
    left = (length - len(res)) // 2
    right = length - left - len(res)

    res = "{}{}{}".format("*" * left, res, "*" * right)

    return res


def equal_codes(code1, code2, ignore_whitespace=True):
    """
    check two codes, if they are equal, return True; else, return False
    :param code1: the code to compare
    :param code2: the other code to compare
    :param ignore_whitespace: ignore whitespace or not, default True
    :return:
    """
    code1 = str(code1)
    code2 = str(code2)

    try:
        whitespace_ptn = "\r|\n|\t|\v| "
        if ignore_whitespace:
            code1 = re.sub(whitespace_ptn, "", code1)
            code2 = re.sub(whitespace_ptn, "", code2)
    except:
        traceback.print_exc()
        print("code1: {}\ncode2: {}".format(code1, code2))

    return code1 == code2


@time_cost
def check_convert_result_between_antlr_and_js(original_tpl_xlsx_path, convert_tpl_xlsx_path, sheet_name=None):
    """
    check the convert result between antlr and js grammar
    :param original_tpl_xlsx_path: the original tpl file, in xlsx
    :param convert_tpl_xlsx_path: the convert tpl file, in xlsx
    :param sheet_name: the specified sheet, None indicates all sheets
    we first read tpl_item list from original_tpl_xlsx_path,
    then read antlr->js->antlr converted tpl_item list from convert_tpl_xlsx_path,
    then use equal_codes() to compare antlr codes in each tpl_item,
    and output result
    :return: dict, in format {SHEETNAME1: [(INTENT1, BRANCH1), ...], ...}
    """
    original_tpl_items = load_jsons_from_xlsx(xlsx_path=original_tpl_xlsx_path, sheet_name=sheet_name)
    convert_tpl_items = load_jsons_from_xlsx(xlsx_path=convert_tpl_xlsx_path, sheet_name=sheet_name)
    diff_info = {}
    for sheet_name in original_tpl_items:
        diff_info[sheet_name] = []
        org_tpl_lst = original_tpl_items[sheet_name]
        cvt_tpl_lst = convert_tpl_items[sheet_name]
        for tpl_item1, tpl_item2 in zip(org_tpl_lst, cvt_tpl_lst):
            if None in tpl_item1:
                tpl_item1.pop(None)
            if None in tpl_item2:
                tpl_item2.pop(None)
            if not equal_codes(tpl_item1, tpl_item2):
                #                 print(tpl_item1)
                #                 print(tpl_item2)
                #                 break
                diff_info[sheet_name].append((tpl_item1["intent"], tpl_item1["param"]))

        if not diff_info[sheet_name]:
            diff_info.pop(sheet_name)

    print(format_string("convert between antlr and js result:"))
    for sheet_name in diff_info:
        print(format_string(sheet_name))
        print("bad convert num: {}".format(len(diff_info[sheet_name])))
        for i, pr in enumerate(diff_info[sheet_name]):
            print("{:03d}: intent: {}, branch: {}".format(i + 1, pr[0], pr[1]))

    return diff_info, original_tpl_items, convert_tpl_items
