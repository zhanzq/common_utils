# !/usr/bin/env python
# encoding=utf-8
# author: zhanzq
# email : zhanzhiqiang09@126.com 
# date  : 2021/12/21
#

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from common_utils.utils import time_cost


def _set_auto_filter(sheet):
    full_range = "A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
    sheet.auto_filter.ref = full_range


def _set_title_format(sheet, size=14, bold=True, color="000000"):
    """
    set title format as bold black font with larger size than default
    :param size: default 14, larger than the default value 12
    :param bold: default True
    :param color: default black
    :return:
    """
    font = Font(name="宋体", size=size, bold=bold, color=color)
    light_green = PatternFill(fill_type='solid', fgColor="55DD55")
    light_blue = PatternFill(fill_type="solid", fgColor="5b94ed")
    for row in sheet.rows:
        for cell in row:
            cell.font = font
            cell.fill = light_blue
        return


def _set_adaptive_column_width(sheet):
    """
    获取每一列的内容的最大宽度
    :param sheet:
    :return:
    """
    i = 0
    col_width = []
    # 每列
    for col in sheet.columns:
        # 每行
        for j in range(len(col)):
            width = len(str(col[j].value).encode("utf-8"))
            if j == 0:
                # 数组增加一个元素
                col_width.append(width)
            else:
                # 获得每列中的内容的最大宽度
                if col_width[i] < width:
                    col_width[i] = width
        i = i + 1

    # 设置列宽
    for i in range(len(col_width)):
        # 根据列的数字返回字母
        col_letter = get_column_letter(i + 1)
        # 当宽度大于100，宽度设置为100
        if col_width[i] > 60:
            sheet.column_dimensions[col_letter].width = 60
        # 只有当宽度大于10，才设置列宽
        elif col_width[i] > 10:
            sheet.column_dimensions[col_letter].width = col_width[i] + 2
        else:
            sheet.column_dimensions[col_letter].width = 10


def _eval_cell(val):
    try:
        val = eval(val)
    except NameError:
        pass
    finally:
        return val


def _convert_pos_26bit(s):
    """
    convert A2 to (2, 1)
    convert AA2 to (2, 27)
    :param s:
    :return:
    """
    first = 0
    second = 0
    for ch in s:
        if 'A' <= ch <= 'Z':
            second *= 26
            second += ord(ch) - ord('A') + 1
        else:
            first *= 10
            first += ord(ch) - ord('0')

    return first, second


def _convert_str_to_cell_range(s):
    """
    convert "A2:A8" to [[(2, 1)], [(3, 1)], [(4, 1)], [(5, 1)], [(6, 1)], [(7, 1)], [(8, 1)]]
    convert "A2:C8" to [[(2, 1), (2, 2), (2, 3)],
                       [(3, 1), (3, 2), (3, 3)],
                       [(4, 1), (4, 2), (4, 3)],
                       [(5, 1), (5, 2), (5, 3)],
                       [(6, 1), (6, 2), (6, 3)],
                       [(7, 1), (7, 2), (7, 3)],
                       [(8, 1), (8, 2), (8, 3)]]
    :param s:
    :return:
    """
    items = s.split(":")
    first1, second1 = _convert_pos_26bit(items[0])
    first2, second2 = _convert_pos_26bit(items[1])

    cell_range = []
    for i in range(first1, first2 + 1):
        row = []
        for j in range(second1, second2 + 1):
            row.append((i, j))
        cell_range.append(row)

    return cell_range


def _get_merged_cells_from_sheet(sheet):
    mp = {}
    if openpyxl.__version__ >= "3.0.1":
        range_lst = sheet.merged_cells.ranges
        for _range in range_lst:
            pos_lst = list(_range.rows)
            i, j = pos_lst[0][0]
            val = sheet[i][j - 1].value  # xlsx row: starts from 0, column: starts from 1
            mp[val] = []
            for lst in pos_lst:
                mp[val].extend(lst)

        return mp
    else:
        range_lst = sheet.merged_cell_ranges
        for range_s in range_lst:
            pos_lst = _convert_str_to_cell_range(range_s)
            i, j = pos_lst[0][0]
            val = sheet[i][j - 1].value  # xlsx row: starts from 0, column: starts from 1
            if val not in mp:
                mp[val] = []
            for lst in pos_lst:
                mp[val].extend(lst)

        return mp


def save_json_list_into_sheet(wb, json_lst, col_name_lst=None, sheet_name="Title", overwrite=True, auto_filter=True):
    """
    save json list into sheet of xlsx file
    :param wb: workbook, fp of xlsx file
    :param json_lst: the data to store
    :param col_name_lst: the column names of the sheet data, ordered according to col_name_lst
    :param sheet_name: the sheet name to store given data
    :param overwrite: overwrite the old data or not, default "True"
    :param auto_filter: auto filter all the fields
    :return:
    """
    if col_name_lst is None:
        col_name_lst = list(json_lst[0].keys())

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if overwrite:
            wb.remove(sheet)
            sheet = wb.create_sheet(title=sheet_name, index=0)
        else:
            return
    else:
        sheet = wb.create_sheet(title=sheet_name, index=0)

    # write first line with col_names
    for j, col_name in enumerate(col_name_lst):
        sheet.cell(row=1, column=j + 1).value = col_name

    # write json obj
    col_sz = len(col_name_lst)
    for i, obj in enumerate(json_lst):
        for j in range(col_sz):
            val = obj.get(col_name_lst[j])
            if val is None:
                val = ""
            if type(val) is float or type(val) is int:
                val = str(val)
                # pass
            else:
                val = str(val)
            sheet.cell(row=i + 2, column=j + 1).value = val  # save the original type
    if auto_filter:
        _set_auto_filter(sheet)

    _set_adaptive_column_width(sheet)
    _set_title_format(sheet)


@time_cost
def save_json_list_into_xlsx(json_lst, xlsx_path, col_name_lst=None, sheet_name: str = "Title",
                             overwrite=True, auto_filter=True):
    """
    write json list into xlsx file, treat keys as the column names
    :param json_lst: data to store with json format
    :param xlsx_path: data path
    :param col_name_lst: the column names of the sheet data, ordered according to col_name_lst
    :param sheet_name: the sheet name to store given data
    :param overwrite: overwrite the old data or not, default "True"
    :param auto_filter: auto filter all the fields
    :return:
    """
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.get_sheet_by_name("Sheet"))

    print("save {} records into sheet: {}".format(len(json_lst), sheet_name))
    if json_lst:
        save_json_list_into_sheet(wb, json_lst=json_lst, col_name_lst=col_name_lst, sheet_name=sheet_name,
                                  overwrite=overwrite, auto_filter=auto_filter)
        wb.save(xlsx_path)


def is_empty_record(item):
    has_valid_val = False
    for key, val in item.items():
        if val:
            has_valid_val = True
            break

    return not has_valid_val


def load_json_list_from_sheet(sheet):
    """
    read data from sheet of xlsx file, and format it in json list
    :param sheet: the sheet name from which to read data
    :return:
    """
    rows = sheet.rows
    col_names = []
    try:
        col_names = [it.value for it in rows.__next__()]
    except StopIteration:
        pass
    json_lst = []
    for row in rows:
        values = []
        for cell in row:
            if cell.font.strike:    # 带有删除线
                values.append(None)
            else:
                values.append(_eval_cell(cell.value))

        json_lst.append(values)
        # values = [_eval_cell(it.value) for it in row]

    merged_cells = _get_merged_cells_from_sheet(sheet)
    for val, pos_lst in merged_cells.items():
        for i, j in pos_lst:
            json_lst[i - 2][j - 1] = val

    out_lst = []
    for item in json_lst:
        obj = {key: val for key, val in zip(col_names, item) if key}
        if not is_empty_record(obj):
            out_lst.append(obj)

    return out_lst


@time_cost
def load_json_list_from_xlsx(xlsx_path, sheet_names=None):
    """
    read data from xlsx file, and format it in dict, keys are all the sheet names or the give sheet_name
    :param xlsx_path: data file path
    :param sheet_names: the sheet names from which to read data
    :return:
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        json_dct = {}
        if not sheet_names:
            valid_sheet_names = wb.sheetnames
        else:
            valid_sheet_names = []
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    valid_sheet_names.append(sheet_name)
                else:
                    print("sheet: {} not found".format(sheet_name))

        for sheet_name in valid_sheet_names:
            sheet = wb[sheet_name]
            json_lst = load_json_list_from_sheet(sheet)
            json_dct[sheet_name] = json_lst
            print("load {} records from sheet: {}".format(len(json_lst), sheet_name))

        return json_dct
    except FileNotFoundError as e:
        print(e)
        return None


def check_title(xlsx_path, necessary_titles, sheet_names=None):
    """
    check the necessary titles are complement or not for all sheets in xlsx file
    :param xlsx_path: input xlsx file path
    :param necessary_titles: necessary titles the sheets must contain
    :param sheet_names: the sheets to check
    :return:
    """
    valid = True
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        json_dct = {}
        if not sheet_names:
            sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            sheet = wb[sheet_name]

            rows = sheet.rows
            col_names = []
            try:
                col_names = [it.value for it in rows.__next__()]
            except StopIteration:
                pass
            for title in necessary_titles:
                if title not in col_names:
                    print("sheet: {} has no title: {}".format(sheet_name, title))
                    valid = False
    except Exception as e:
        print(e)
    finally:
        return valid


def _test():
    xlsx_path = "./test.xlsx"
    json_dct = load_json_list_from_xlsx(xlsx_path=xlsx_path)
    for sheet_name, json_lst in json_dct.items():
        save_json_list_into_xlsx(xlsx_path=xlsx_path, sheet_name=sheet_name, json_lst=json_lst)
    print(json_dct.keys())


if __name__ == "__main__":
    _test()
